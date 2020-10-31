import shutil
import string
from socket import socket

from selenium import webdriver
from selenium.webdriver.support.ui import Select
import time
import urllib.request
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows

#Start Settings,Brand Name,Mode,Criteria,Generic Title


 # try:
driver = webdriver.Chrome(executable_path="C:/Users/Mian/Desktop/project/chromedriver.exe")
driver.get("https://webcat.zf.com/index.asp?KAT_KZ=PNBTMFA&ARTNR=803+054&EINSPNR_ART=32&KMODNR=0&LASTPAGE=nc2_teile_info.asp")
    # error = driver.find_element_by_class_name("error-code")
 #    if (error.text=='DNS_PROBE_FINISHED_NO_INTERNET') or (error.text=='ERR_NAME_NOT_RESOLVED'):
 #        print("No Internet")
 # except OSError:
 #    print("Error !!!!")


print("Confirm to Start the Process")
val = input("Enter your value: ")
if (val == 'Start'):
        print("Starting the Process")


# Global variables
model = '';
brand = '';
ZF = '';
MFN = '';
bin = '';
count='';
model=''
brand=''

#Intro function to get introduction
def intro(trb):
    try:
        criteria = '';
        rfarr=[]
        print("Printing Introduction Information")
        # Get Model and Brand
        global count, brand, model
        brand = trb[0].find_elements_by_tag_name('li')[0]
        model = trb[0].find_elements_by_tag_name('li')[1]
        generic = trb[0].find_elements_by_tag_name('li')[2]
        generic = generic.text.strip()
        model = model.text.replace(" ", "")
        brand = brand.text.replace("brand: ", "")
        count = len(generic)
        # (count)
        # write Criteria Excel
        criteriaex = []
        criteriaex.append(brand)
        criteriaex.append(model)
        criteriaex.append(generic)

        print(model)
        # Get Criteria
        criteriatr = driver.find_element_by_xpath('/html/body/table[2]/tbody/tr/td[1]/table/tbody/tr[8]')
        criteriali = criteriatr.find_elements_by_tag_name('li')
        criteriabool = True
        for li in criteriali:
            if criteriabool:
                li = li.text.strip()
                criteria = criteria + li
                criteriabool = False
            else:
                li = li.text.strip()
                criteria = criteria + " // " + li

        # Writing to the refernce excel sheet
        rfarr.append(brand)
        rfarr.append(model)
        referncewr.append(rfarr)
        rfarr.clear()

        #Writing To the criteria  excel sheet
        criteriaex.append(criteria)
        criteriast.append(criteriaex)
        criteriaex.clear()
    except Exception as e:
        print(" ? ? Exception in Introduction ? ? ")
        print(e)

# Printing ZF Number

def zf(trb):
    print("Printing ZF Number")
    try:
        # ZF Aftermarket compare numbers
        zftables = trb[indextrb - 2].find_elements_by_tag_name('table')
        zfuls = trb[indextrb - 2].find_elements_by_tag_name('ul')
        counter1 = 0

        # write ZF Excel
        zfex = []

        for zftable in zftables:
            zfheading = zftable.find_element_by_tag_name('td').text
            # print(zfheading)
            zflis = zfuls[counter1].find_elements_by_tag_name('li')
            for zfli in zflis:
                zfli = zfli.text
                # print(zfli)
                zfex.append(brand)
                zfex.append(model)
                zfex.append(zfheading)
                zfex.append(zfli)
                zfst.append(zfex)
                zfex.clear()
            counter1 = counter1 + 1
    except Exception as e:
        print(" ? ? Exception In ZF Number ? ? ")
        print(e)


# Printing Manufacture Number
def mn(trb):
        print("Printing Manufacture Number")


        # write manufacture Excel
        mnex = []
        try:
            mntables = trb[indextrb - 3].find_elements_by_tag_name('table')
            mnuls = trb[indextrb - 3].find_elements_by_tag_name('ul')
            counter1 = 0
            for mntable in mntables:
                mnheading = mntable.find_element_by_tag_name('td').text
                # print(mnheading)
                mnlis = mnuls[counter1].find_elements_by_tag_name('li')
                for mnli in mnlis:
                    mnli = mnli.text
                    # print(mnli)
                    mnex.append(brand)
                    mnex.append(model)
                    mnex.append(mnheading)
                    mnex.append(mnli)
                    mnst.append(mnex)
                    mnex.clear()
                counter1 = counter1 + 1
        except Exception as e:
            print(" ? ? Exception for writing Manufacture Numbers ? ? ")
            print(e)


# Printing Built In Number
def built():

     try:
        print("Printing Built-in Number")
        builtintb = driver.find_element_by_xpath('/html/body/table[2]/tbody/tr/td[3]/table')
        trows2 = builtintb.find_elements_by_tag_name('tr')
        child = False

        binex = []
        for index, trow2 in enumerate(trows2):
            if index + 1 > 3:
                if trow2.get_attribute('class') == 'TableRow2':
                    trowcntnt = trow2.text
                    trowcntnt = trowcntnt.replace(" :", "")
                    child = True
                    continue
                elif child:
                    childcntnt = trow2.text
                    childcntnt = childcntnt.strip()
                    child = False
                    continue
                else:
                    subchildcntntbs = trow2.find_elements_by_tag_name('b')
                    subchildcntntss = trow2.find_elements_by_tag_name('span')
                    for index2, subchildcntntb in enumerate(subchildcntntbs):

                        if subchildcntntss[index2].text:
                            subchildcntnt = subchildcntntb.text + subchildcntntss[index2].text
                        else:
                          subchildcntnt = subchildcntntb.text

                        subchildcntnt = subchildcntnt.strip()
                        binex.append(brand)
                        binex.append(model)
                        binex.append(trowcntnt)
                        binex.append(childcntnt)
                        binex.append(subchildcntnt)
                        # print(trowcntnt)
                        # print(childcntnt)
                        # print(subchildcntnt)
                        bist.append(binex)
                        binex.clear()

                    child = True
     except Exception as e:

         print(" ? ? Exception in Built In Number ? ? ")
         print('e')


# Load existing workbook
filename = 'C:/Users/Mian/Desktop/project/sample.xlsx'
filenameread = 'C:/Users/Mian/Desktop/project/existing.xlsx'
wb = load_workbook(filename)
rf = load_workbook(filenameread)

# Get Existing Sheet
refernceex = rf['reference number']

# Worksheet to write data in
referncewr=wb['reference number']
criteriast = wb['criteria']
mnst = wb['ZF Aftermarket compare numbers']
zfst = wb['manufacture compare number']
bist = wb['built-in']




#for Loop
for i in range(2591,3799):

        print("===Starting The Process===")
        print("Scraping Url # " + str(i-1))
        # This is a sample Python script.

        # Press Shift+F10 to execute it or replace it with your code.
        # Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

        a3 = refernceex.cell(row=i, column=3)
        try:
           driver.get(a3.value)
        except:
         print(" Error in Getting Url")
         time.sleep(10)
         try:
            driver.get(a3.value)
         except:
            print(" Error Again in getting url 1")
            time.sleep(100)
            try:
                driver.get(a3.value)
            except:
                print(" Error Again in getting url 2")
                time.sleep(200)
                try:
                    driver.get(a3.value)
                except:
                    print(" Error Again in getting url 3")
                    time.sleep(500)
                    try:
                        driver.get(a3.value)
                    except:
                        print(" Error Again in getting url 4")
                        time.sleep(1000)
                        driver.get(a3.value)
                        wb.save(filename)




        driver.switch_to.frame("NETCAT")
        trb = driver.find_elements_by_class_name("TableRowBlank")
        # print(trb)
        indextrb = len(trb)

        #call the function
        intro(trb)
        zf(trb)
        mn(trb)
        built()


        try:
            img = driver.find_element_by_class_name('img_preview')
            imga = img.find_element_by_tag_name('a')
            imghref = imga.get_attribute('href')
            # print(imghref)
            driver.maximize_window()
        except:
             print("Image Not Found")
        try:
            driver.get(imghref)



        except:

            print("Error In getting Images 1")
            time.sleep(10)
            try:
             driver.get(imghref)
            except:

                print("Error In getting Images 1")
                time.sleep(200)
                try:
                    driver.get(imghref)
                except:

                    print("Error In getting Images 2")
                    time.sleep(500)
                    try:
                        driver.get(imghref)
                    except:
                        print("Error in getting Images 3")
                        time.sleep(1000)
                        driver.get(imghref)
                        wb.save(filename)
        try:
            driver.switch_to.frame("NETCAT")
            downloadimgclass = driver.find_element_by_class_name("img_preview")
        # driver.switch_to.frame("NETCAT")
        # driver.get_screenshot_as_png()
        # img=driver.find_element_by_class_name('img_preview')
            imga=downloadimgclass.find_element_by_tag_name('img')
        # imghref=imga.get_attribute('src')
        except:
            continue


        # V
        try:
            import pyperclip, pyautogui

            #  Move to the specified location, right click
            count = int(count) * 5
            pyautogui.rightClick(x=400 + count, y=270)
            pyautogui.typewrite(['V'])

        # Copy address and file name
            pyperclip.copy(model)

        # Wait for the window to open, so as not to conflict with the command, the paste fails, and tried it many times before it has 0.8.
            time.sleep(0.8)
        #
            pyautogui.hotkey('ctrlleft', 'V')
        #  can also be pasted
        # pyautogui.keyDown('ctrl')
        # pyautogui.press('v')
        # pyautogui.keyUp('ctrl')

        #
            pyautogui.press('enter')
            pyautogui.press('left')
            pyautogui.press('enter')
        except:
            print("Error in processing")

        # Save Excel File
        if i%10==0:
         wb.save(filename)
        # Confirm the Process



    # except Exception as e:
    #     print(e)
    #     time.sleep(5)
    #     continue


